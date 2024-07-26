import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re
import os

#comprovar si existe el archivo
file = 'datos.xlsx'
if os.path.exists(file):
    wb = load_workbook(file)
    ws = wb.active
else:
#crear el libro
    wb = Workbook()
    ws = wb.active
    ws.append(["Nombre", "Edad", "Email", "Teléfono", "Dirección"])

def guardar_datos():
    nombre = entry_nombre.get()
    edad = entry_edad.get()
    email =entry_email.get()
    telefono = entry_telefono.get()
    direccion =entry_direccion.get()
        #validacion de que se llenen todos los campos
    if not nombre or not edad or not email or not telefono or not direccion:
        messagebox.showwarning(title="Advertencia", message="Todos los campos son obligatorios")
        return
    try:  # validacion con try de que telefono y edad son campos int
        edad = int(edad)
        telefono = int(telefono)
    except ValueError:
        messagebox.showwarning(title="Advertencia", message="Edad y Telefono son campos numericos")
        return


    # validacion del formato email
    pattern = r"^[^@]+@[^@]+\.[^@]+$"

    if not re.match(pattern, email): 
        messagebox.showwarning(title="Advertencia", message="El correo no es valido")
        return

    ws.append([nombre, edad, email, telefono, direccion])
    #file es la variable que contiene el nombre del archivo")
    wb.save(file)
    messagebox.showwarning(title="Información", message="Datos guardados con éxito")

    #borrar los datos de las cajas de textos
    entry_nombre.delete(0, tk.END)
    entry_edad.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)
    entry_direccion.delete(0, tk.END)


#crear ventana
root = tk.Tk()
root.geometry("300x250+560+240")
root.title("Formulario de Entrada de Datos")
root.configure(background="#4B6587")
label_style = {"bg":"#4B6587", "fg":"white"}
entry_style = {"bg":"#D3D3D3", "fg":"black"}

label_nombre = tk.Label(root, text="Nombre", **label_style)
label_nombre.grid(row=0, column=0, padx=10, pady=5)
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=10, pady=5)

label_edad= tk.Label(root, text="Edad", **label_style)
label_edad.grid(row=1, column=0, padx=10, pady=5)
entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=1, column=1, padx=10, pady=5)

label_email = tk.Label(root, text="Email", **label_style)
label_email.grid(row=2, column=0, padx=10, pady=5)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=2, column=1, padx=10, pady=5)

label_telefono = tk.Label(root, text="Telefono", **label_style)
label_telefono.grid(row=3, column=0, padx=10, pady=5)
entry_telefono = tk.Entry(root, **entry_style)
entry_telefono.grid(row=3, column=1, padx=10, pady=5)

label_direccion = tk.Label(root, text="Dirección", **label_style)
label_direccion.grid(row=4, column=0, padx=10, pady=5)
entry_direccion = tk.Entry(root, **entry_style)
entry_direccion.grid(row=4, column=1, padx=10, pady=5)

boton_guardar = tk.Button(root, text="Guardar", command=guardar_datos, bg='#6D8299',fg='#6D8299', relief='flat', borderwidth=1)
boton_guardar.grid(row=5, column=0, columnspan=2, padx=10, pady=10)

root.mainloop()


# darle nombre y guardar archivo
#file = "datos.xlsx"
#wb.save(file)